import json
import numpy as np
import os
import numpy as np
import matplotlib.pyplot as plt
import datetime
from classes.packet_registry import packet_registry as registry

class MetricsManager:
    def __init__(self):
        self.metrics = {}

    def initialize(self, max_hops, topology_file, functions_sequence, mean_interval_ms, reconnect_interval_ms, disconnect_probability, algorithms):
        """Inicializa las métricas para una nueva simulación con múltiples algoritmos."""
        self.metrics = {
            "simulation_id": 1,
            "parameters": {
                "max_hops": max_hops,
                "algorithms": algorithms,  # Ahora guardamos una lista de algoritmos
                "mean_interval_ms": mean_interval_ms,
                "reconnect_interval_ms": reconnect_interval_ms,
                "topology_file": topology_file,
                "functions_sequence": [func.value for func in functions_sequence],
                "disconnect_probability": disconnect_probability
            },
            "total_time": None,
            "runned_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

        # Inicializamos métricas para cada algoritmo seleccionado
        for algorithm in algorithms:
            self.metrics[algorithm] = {
                "success_rate": 0.0,
                "episodes": [],
                "penalty": None if algorithm != "Q_ROUTING" else 0.0
            }

    def log_episode(self, algorithm, episode_number, start_time, end_time, episode_success, route, total_hops, dynamic_changes):
        """Registra las métricas de un episodio en ejecución para un algoritmo específico."""
        if algorithm not in self.metrics:
            self.metrics[algorithm] = {
                "success_rate": 0.0,
                "episodes": []
            }

        self.metrics[algorithm]["episodes"].append({
            "episode_number": episode_number,
            "start_time": start_time,
            "end_time": end_time,
            "episode_duration": end_time - start_time,
            "episode_success": episode_success,
            "route": route,
            "total_hops": total_hops,
            "dynamic_changes": dynamic_changes,
            "dynamic_changes_count": len(dynamic_changes)
        })

    def finalize_simulation(self, total_time, successful_episodes, episodes):
        """Finaliza la simulación y guarda los resultados."""
        algorithm = self.metrics["parameters"]["algorithms"][-1]
        self.metrics[algorithm]["success_rate"] = successful_episodes / episodes if episodes > 0 else 0.0
        self.metrics["total_time"] = total_time
        self.save_metrics_to_file()
        self.save_results_to_excel()
        # self.generar_comparative_graphs_from_excel()

    def save_metrics_to_file(self, directory="../results/single-run"):
        """
        Guarda las métricas de la simulación en un archivo JSON con el simulation_id en el nombre.
        Crea la carpeta `../results/simulations` si no existe.
        """
        os.makedirs(directory, exist_ok=True)

        filename = f"{directory}/simulation_{self.metrics['simulation_id']}.json"

        try:
            with open(filename, "w", encoding="utf-8") as file:
                json.dump(self.metrics, file, indent=4)
            print(f"\nMétricas de la simulación guardadas en {filename}")
        except Exception as e:
            print(f"\nError al guardar las métricas: {e}")

    def save_results_to_excel(self, filename="../results/resultados_simulacion.xlsx"):
        """
        Guarda los datos de la simulación en un archivo Excel, con una hoja por algoritmo.
        Incluye el `packet_log` completo para cada episodio para debug.
        """
        import os
        import json
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        os.makedirs("../results", exist_ok=True)

        # Verificar si el archivo es corrupto antes de continuar
        if os.path.exists(filename):
            try:
                pd.ExcelFile(filename)
            except Exception:
                print(f"Archivo corrupto detectado: {filename}. Eliminando y regenerando...")
                os.remove(filename)

        # Validar que self.metrics tiene datos
        if not self.metrics:
            print("⚠️ No hay datos en self.metrics. No se guardará el Excel.")
            return

        # Definir estructura de almacenamiento de métricas
        metrics_data = {
            algorithm: {
                "episode": [],
                "start_time": [],
                "end_time": [],
                "episode_duration": [],
                "episode_success": [],
                "total_hops": [],
                "dynamic_changes": [],
                "packet_log_raw": [],  # Guardar packet log en JSON como string
            }
            for algorithm in self.metrics.keys() if algorithm not in ["simulation_id", "parameters", "total_time", "runned_at"]
        }
        print(f"\nAlgoritmos encontrados en las métricas: {list(metrics_data.keys())}")
        # Recorrer episodios y almacenar métricas
        for algorithm, episodes in self.metrics.items():
            if algorithm in ["simulation_id", "parameters", "total_time", "runned_at"]:
                continue
            for episode_data in episodes["episodes"]:
                print(f"ℹ️ Procesando episodio {episode_data['episode_number']} para {algorithm}...")
                episode_number = episode_data["episode_number"]
                # Obtener packet log del episodio
                packet_log = registry.packet_log.get(episode_number, {})

                # Agregar datos a las listas de métricas
                metrics_data[algorithm]["episode"].append(episode_number)
                metrics_data[algorithm]["start_time"].append(episode_data.get("start_time", 0))
                metrics_data[algorithm]["end_time"].append(episode_data.get("end_time", 0))
                metrics_data[algorithm]["episode_duration"].append(episode_data.get("episode_duration", 0))
                metrics_data[algorithm]["episode_success"].append(episode_data.get("episode_success", False))
                metrics_data[algorithm]["total_hops"].append(episode_data.get("total_hops", 0))
                metrics_data[algorithm]["dynamic_changes"].append(len(episode_data.get("dynamic_changes", [])))

                # Guardar el packet log en JSON string (para debug)
                try:
                    packet_log_json = json.dumps(packet_log, indent=2)
                except Exception as e:
                    print(f"⚠️ Error serializando packet_log para el episodio {episode_number}: {e}")
                    packet_log_json = "{}"

                metrics_data[algorithm]["packet_log_raw"].append(packet_log_json)

        # Guardar en Excel
        with pd.ExcelWriter(filename, engine="openpyxl", mode="w") as writer:
            for algorithm, data in metrics_data.items():
                if not data["episode"]:  # Si no hay episodios, no crear la hoja
                    print(f"⚠️ No hay episodios registrados para el algoritmo {algorithm}.")
                    continue

                df = pd.DataFrame(data)
                df.to_excel(writer, sheet_name=algorithm, index=False)

        # Ajustar ancho de columnas para mejorar legibilidad
        wb = load_workbook(filename)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for column in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
                ws.column_dimensions[get_column_letter(column[0].column)].width = max_length + 2
        wb.save(filename)

        print(f"\n✅ Resultados guardados en {filename}.")

    def generar_comparative_graphs_from_excel(self, filename="../results/resultados_simulacion.xlsx"):
        """
        Genera gráficos comparativos basados en las métricas de la simulación, comparando todos los algoritmos en un solo gráfico por métrica.
        """
        import matplotlib.pyplot as plt
        import pandas as pd
        import os

        os.makedirs("../results", exist_ok=True)
        xls = pd.ExcelFile(filename)

        # Diccionario para almacenar datos de todas las hojas
        all_data = {
            "episode_duration": {},
            "hops_promedio": {},
            "total_hops": {},
            "average_delivery_time": {},
            "success_rate": {},
            "episode_success": {}
        }

        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            all_data["episode_duration"][sheet_name] = df["episode_duration"]
            all_data["hops_promedio"][sheet_name] = df["total_hops"] / df["episode"]
            all_data["total_hops"][sheet_name] = df["total_hops"]
            all_data["average_delivery_time"][sheet_name] = df["episode_duration"] / df["total_hops"]
            all_data["success_rate"][sheet_name] = df["episode_success"].fillna(False).infer_objects(copy=False)
            all_data["episode_success"][sheet_name] = df["episode_success"]

        # Gráfico comparativo de Duración del Episodio
        plt.figure(figsize=(10, 6))
        for algorithm, data in all_data["episode_duration"].items():
            plt.plot(data, label=f"{algorithm}")
        plt.title("Comparación de Duración del Episodio entre Algoritmos")
        plt.xlabel("Episodio")
        plt.ylabel("Duración (ms)")
        plt.grid()
        plt.legend()
        plt.savefig("../results/Comparacion_Duracion_Episodio.png")
        plt.close()

        # Gráfico comparativo de Hops Promedio
        plt.figure(figsize=(10, 6))
        for algorithm, data in all_data["hops_promedio"].items():
            plt.plot(data, label=f"{algorithm}")
        plt.title("Comparación de Hops Promedio entre Algoritmos")
        plt.xlabel("Episodio")
        plt.ylabel("Hops Promedio")
        plt.grid()
        plt.legend()
        plt.savefig("../results/Comparacion_Hops_Promedio.png")
        plt.close()

        # Gráfico comparativo de Tiempo Promedio de Entrega
        plt.figure(figsize=(10, 6))
        for algorithm, data in all_data["average_delivery_time"].items():
            plt.plot(data, label=f"{algorithm}")
        plt.title("Comparación de Tiempo Promedio de Entrega entre Algoritmos")
        plt.xlabel("Episodio")
        plt.ylabel("Tiempo Promedio de Entrega (ms)")
        plt.grid()
        plt.legend()
        plt.savefig("../results/Comparacion_Tiempo_Promedio_Entrega.png")
        plt.close()

        # Gráfico comparativo de Tasa de Éxito
        plt.figure(figsize=(10, 6))
        for algorithm, data in all_data["success_rate"].items():
            plt.plot(data, label=f"{algorithm}")
        plt.title("Comparación de Tasa de Éxito entre Algoritmos")
        plt.xlabel("Episodio")
        plt.ylabel("Tasa de Éxito")
        plt.grid()
        plt.legend()
        plt.savefig("../results/Comparacion_Tasa_Exito.png")
        plt.close()

        # Gráfico de columnas verticales de Tasa de Éxito por Episodio
        plt.figure(figsize=(10, 6))

        # Initialize dictionaries to count TRUE, FALSE, and blank values for each algorithm
        success_counts = {algorithm: {"TRUE": 0, "FALSE": 0, "BLANK": 0} for algorithm in all_data["episode_success"].keys()}

        # Count the occurrences of TRUE, FALSE, and blank values for each algorithm
        for algorithm, data in all_data["episode_success"].items():
            for value in data:
                if pd.isna(value):
                    success_counts[algorithm]["BLANK"] += 1
                elif value:
                    success_counts[algorithm]["TRUE"] += 1
                else:
                    success_counts[algorithm]["FALSE"] += 1

        # Plot the counts for each algorithm
        bar_width = 0.25
        index = np.arange(len(success_counts))

        # Plot TRUE values
        plt.bar(index, [success_counts[alg]["TRUE"] for alg in success_counts], bar_width, label="TRUE")

        # Plot FALSE values
        plt.bar(index + bar_width, [success_counts[alg]["FALSE"] for alg in success_counts], bar_width, label="FALSE")

        # Plot Blank values
        plt.bar(index + 2 * bar_width, [success_counts[alg]["BLANK"] for alg in success_counts], bar_width, label="BLANK")

        plt.title("Tasa de Éxito por Episodio entre Algoritmos")
        plt.xlabel("Algoritmo")
        plt.ylabel("Cantidad")
        plt.xticks(index + bar_width, success_counts.keys())
        plt.grid()
        plt.legend()
        plt.savefig("../results/Tasa_Exito_Columnas.png")
        plt.close()

        print("\nGráficos comparativos generados en '../results/'.")


    def generate_heat_map(self, q_tables):
        q_table_data = []
        for q_table in q_tables:
            for state, actions in q_table.items():
                for action, q_value in actions.items():
                    q_table_data.append((state, action, q_value))

        if not q_table_data:
            print(f"No Q-table data available.")
            return

        # Extract unique states and actions
        states = sorted(set(state for state, _, _ in q_table_data))
        actions = sorted(set(action for _, action, _ in q_table_data))

        # Create a matrix to hold Q-values
        q_matrix = np.zeros((len(states), len(actions)))

        for state, action, q_value in q_table_data:
            state_index = states.index(state)
            action_index = actions.index(action)
            q_matrix[state_index, action_index] = q_value

        fig, ax = plt.subplots()
        cax = ax.imshow(q_matrix, cmap='RdYlGn', aspect='auto', vmin=0, vmax=1)

        # Add color bar
        fig.colorbar(cax)

        # Set axis labels
        ax.set_xticks(np.arange(len(actions)))
        ax.set_yticks(np.arange(len(states)))
        ax.set_xticklabels(actions)
        ax.set_yticklabels(states)

        # Annotate each cell with its value
        for i in range(len(states)):
            for j in range(len(actions)):
                ax.text(j, i, f'{q_matrix[i, j]:.2f}', ha='center', va='center', color='black')

        plt.xlabel('Actions (Next Node ID)')
        plt.ylabel('States (Node ID)')
        plt.title(f'Q-Table Heat Map')

        output_folder = '../results'
        # Save the heat map as a .png file
        filename = os.path.join(output_folder, f'q_table_heat_map.png')
        plt.savefig(filename)
        plt.close()

        print(f'Heat map saved to {filename}')
