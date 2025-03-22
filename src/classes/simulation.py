import json
import numpy as np
import os
import time
import numpy as np
from visualization import generate_heat_map, print_q_table
from tabulate import tabulate
import pandas as pd
import matplotlib.pyplot as plt
import threading
import datetime
from classes.base import EpisodeEnded

class Simulation:
    def __init__(self, network, sender_node):
        self.network = network
        self.clock = 0  # Tiempo en milisegundos
        self.running = False  # Control del reloj
        self.lock = threading.Lock()  # Para sincronizar accesos al reloj
        self.sender_node = sender_node
        self.max_hops = None
        self.metrics = {
            "simulation_id": 1,
            "parameters": {
                "max_hops": None,
                "algorithms": [],  # Se llenará en `start()`
                "mean_interval_ms": None,
                "reconnect_interval_ms": None,
                "topology_file": None,
                "functions_sequence": None
            },
            "total_time": None,
            "runned_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "DIJKSTRA": {"success_rate": None, "episodes": []},
            "BELLMAN_FORD": {"success_rate": None, "episodes": []},
            "Q_ROUTING": {"success_rate": None, "penalty": None, "episodes": []}
        }

    def set_max_hops(self, max_hops):
        self.max_hops = max_hops

    def set_mean_interval_ms(self, mean_interval_ms):
        self.mean_interval_ms = mean_interval_ms

    def set_topology_file(self, topology_file):
        self.topology_file = topology_file

    def set_functions_sequence(self, functions_sequence):
        self.functions_sequence = functions_sequence

    def start_clock(self):
        """Inicia un hilo dedicado al reloj central."""
        self.running = True
        threading.Thread(target=self._run_clock, daemon=True).start()

    def stop_clock(self):
        """Detiene el hilo del reloj."""
        self.running = False

    def _run_clock(self):
        """Incrementa el reloj centralizado continuamente en milisegundos."""
        while self.running:
            with self.lock:
                self.clock += 1
            time.sleep(0.001)  # 1 ms en tiempo real

    def get_current_time(self):
        """Obtiene el tiempo actual del reloj central."""
        with self.lock:
            return self.clock

    def tick(self):
        """Avanza el reloj centralizado."""
        self.clock += self.time_increment
        return self.clock

    def get_current_time(self):
        """Devuelve el tiempo actual del reloj central."""
        return self.clock

    def reset_simulation(self):
        self.metrics["simulation_id"] = self.metrics["simulation_id"] + 1

        self.metrics = {
            "parameters": {
                "max_hops": None,
                "algorithms": [],
                "mean_interval_ms": None,
                "reconnect_interval_ms": None,
                "topology_file": None,
                "functions_sequence": None,
                "disconnect_probability": None
            },
            "total_time": None,
            "runned_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "DIJKSTRA": {"success_rate": None, "episodes": []},
            "BELLMAN_FORD": {"success_rate": None, "episodes": []},
            "Q_ROUTING": {"success_rate": None, "penalty": None, "episodes": []}
        }

    def start(self, algorithm_enum, episodes, functions_sequence, mean_interval_ms, reconnect_interval_ms, topology_file, penalty, disconnect_probability):
        """
        Inicia la simulación con el algoritmo seleccionado y registra métricas basadas en tiempo.
        """
        algorithm = algorithm_enum.name
        self.network.set_simulation_clock(self)
        self.start_clock()
        self.network.start_dynamic_changes()

        if algorithm_enum.name not in self.metrics["parameters"]["algorithms"]:
            self.metrics["parameters"]["algorithms"].append(algorithm_enum.name)

        # **Inicializar estructura global de métricas**
        self.metrics["parameters"]["max_hops"] = self.max_hops
        self.metrics["parameters"]["mean_interval_ms"] = mean_interval_ms
        self.metrics["parameters"]["reconnect_interval_ms"] = reconnect_interval_ms
        self.metrics["parameters"]["topology_file"] = topology_file
        self.metrics["parameters"]["functions_sequence"] = [func.value for func in functions_sequence]
        self.metrics["parameters"]["disconnect_probability"] = disconnect_probability

        # **Inicializar estructura de métricas por algoritmo si no existe**
        if algorithm not in self.metrics:
            self.metrics["algorithms"].append(algorithm)
            self.metrics[algorithm] = {
                "success_rate": 0.0,
                "episodes": [],  # Aquí inicializamos correctamente la lista de episodios
                "penalty": penalty if algorithm == "Q_ROUTING" else None  # Agregar penalty solo en Q-Routing
            }

        if algorithm == "Q_ROUTING":
            self.metrics[algorithm]["penalty"] = penalty if penalty else 0.0

        successful_episodes = 0  # Contador de episodios exitosos

        # **Iterar sobre los episodios**
        for episode_number in range(1, episodes + 1):
            print(f'\n\n=== Starting Episode #{episode_number} ({algorithm}) ===\n')

            start_time = self.get_current_time()

            # **Estado inicial de la red**
            node_info = [
                [node.node_id, node.status, node.lifetime, node.reconnect_time]
                for node in self.network.nodes.values()
            ]
            headers = ["Node ID", "Connected", "Lifetime", "Reconnect Time"]
            print(tabulate(node_info, headers=headers, tablefmt="grid"))

            # **Ejecutar episodio**
            if algorithm == "Q_ROUTING":
                try:
                    self.sender_node.start_episode(episode_number, self.max_hops, functions_sequence, penalty)
                except EpisodeEnded:
                    print(f'\n\n=== Episode #{episode_number} ended ===\n')
            else:
                try:
                    self.sender_node.start_episode(episode_number, self.max_hops, functions_sequence, False)
                except EpisodeEnded:
                    print(f'\n\n=== Episode #{episode_number} ended ===\n')

            end_time = self.get_current_time()
            episode_duration = end_time - start_time

            # **Recolección de datos**
            if episode_number in self.network.packet_log:
                self.network.packet_log[episode_number]["episode_duration"] = episode_duration

            episode_data = self.network.packet_log.get(episode_number, {})
            episode_success = episode_data.get("episode_success", False)
            route = episode_data.get("route", [])
            total_hops = len(route)
            dynamic_changes = self.network.get_dynamic_changes_by_episode(start_time, end_time)

            # **Guardar métricas del episodio**
            self.metrics[algorithm]["episodes"].append({
                "episode_number": episode_number,
                "start_time": start_time,
                "end_time": end_time,
                "episode_duration": episode_duration,
                "episode_success": episode_success,
                "route": route,
                "total_hops": total_hops,
                "dynamic_changes": dynamic_changes,
                "dynamic_changes_count": len(dynamic_changes)
            })

            if episode_success:
                successful_episodes += 1  # Contar episodios exitosos

            # **Mostrar métricas en consola**
            print(f"\n Episode #{episode_number} Metrics ({algorithm}):")
            print(f"  - Duración total del episodio: {episode_duration} ms")
            print(f"  - Episodio {'Éxito' if episode_success else 'Fallo'}")
            print(f"  - Ruta seguida: {route}")
            print(f"  - Hops efectivos: {total_hops}")
            print(f"  - Cambios dinámicos en este episodio: {len(dynamic_changes)}")

        # **Calcular tasa de éxito del algoritmo**
        self.metrics[algorithm]["success_rate"] = successful_episodes / episodes if episodes > 0 else 0.0

        # **Detener reloj al finalizar la simulación**
        self.stop_clock()
        self.network.stop_dynamic_changes()
        self.network.packet_log = {}
        self.metrics["total_time"] = self.get_current_time()

        print("\n[Simulation] Simulation finished and clock stopped.")

        # print(json.dumps(self.metrics, indent=4))

        self.save_metrics_to_file()
        self.save_results_to_excel()
        # self.generar_individual_graphs_from_excel()
        self.generar_comparative_graphs_from_excel()

        q_tables = []
        if algorithm == "Q_ROUTING":
            for node in self.network.nodes.values():
                q_tables.append(node.application.q_table)
        self.generate_heat_map(q_tables)

    def save_metrics_to_file(self, directory="../results/single-run"):
        """
        Guarda las métricas de la simulación en un archivo JSON con el simulation_id en el nombre.
        Crea la carpeta `../results/simulations` si no existe.
        """
        os.makedirs(directory, exist_ok=True)  # Crear directorio si no existe

        filename = f"{directory}/simulation_{self.metrics['simulation_id']}.json"

        try:
            with open(filename, "w", encoding="utf-8") as file:
                json.dump(self.metrics, file, indent=4)
            print(f"\nMétricas de la simulación guardadas en {filename}")
        except Exception as e:
            print(f"\nError al guardar las métricas: {e}")

    def save_results_to_excel(self, filename="../results/resultados_simulacion.xlsx"):
        """
        Guarda los datos de la simulación en un archivo Excel, con una hoja por algoritmo.
        Incluye el `packet_log` completo para cada episodio para debug.
        """
        import os
        import json
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        os.makedirs("../results", exist_ok=True)

        # Verificar si el archivo es corrupto antes de continuar
        if os.path.exists(filename):
            try:
                pd.ExcelFile(filename)
            except Exception:
                print(f"⚠️ Archivo corrupto detectado: {filename}. Eliminando y regenerando...")
                os.remove(filename)

        # Validar que self.metrics tiene datos
        if not self.metrics:
            print("⚠️ No hay datos en self.metrics. No se guardará el Excel.")
            return

        # Definir estructura de almacenamiento de métricas
        metrics_data = {
            algorithm: {
                "episode": [],
                "start_time": [],
                "end_time": [],
                "episode_duration": [],
                "episode_success": [],
                "total_hops": [],
                "dynamic_changes": [],
                "packet_log_raw": [],  # Guardar packet log en JSON como string
            }
            for algorithm in self.metrics.keys() if algorithm not in ["simulation_id", "parameters", "total_time", "runned_at"]
        }
        print(f"\nℹ️ Algoritmos encontrados en las métricas: {list(metrics_data.keys())}")
        # Recorrer episodios y almacenar métricas
        for algorithm, episodes in self.metrics.items():
            if algorithm in ["simulation_id", "parameters", "total_time", "runned_at"]:
                continue
            for episode_data in episodes["episodes"]:
                print(f"ℹ️ Procesando episodio {episode_data['episode_number']} para {algorithm}...")
                episode_number = episode_data["episode_number"]
                # Obtener packet log del episodio
                packet_log = self.network.packet_log.get(episode_number, {})

                # Agregar datos a las listas de métricas
                metrics_data[algorithm]["episode"].append(episode_number)
                metrics_data[algorithm]["start_time"].append(episode_data.get("start_time", 0))
                metrics_data[algorithm]["end_time"].append(episode_data.get("end_time", 0))
                metrics_data[algorithm]["episode_duration"].append(episode_data.get("episode_duration", 0))
                metrics_data[algorithm]["episode_success"].append(episode_data.get("episode_success", False))
                metrics_data[algorithm]["total_hops"].append(episode_data.get("total_hops", 0))
                metrics_data[algorithm]["dynamic_changes"].append(len(episode_data.get("dynamic_changes", [])))

                # Guardar el packet log en JSON string (para debug)
                try:
                    packet_log_json = json.dumps(packet_log, indent=2)
                except Exception as e:
                    print(f"⚠️ Error serializando packet_log para el episodio {episode_number}: {e}")
                    packet_log_json = "{}"

                metrics_data[algorithm]["packet_log_raw"].append(packet_log_json)

        # Guardar en Excel
        with pd.ExcelWriter(filename, engine="openpyxl", mode="w") as writer:
            for algorithm, data in metrics_data.items():
                if not data["episode"]:  # Si no hay episodios, no crear la hoja
                    print(f"⚠️ No hay episodios registrados para el algoritmo {algorithm}.")
                    continue

                df = pd.DataFrame(data)
                df.to_excel(writer, sheet_name=algorithm, index=False)

        # Ajustar ancho de columnas para mejorar legibilidad
        wb = load_workbook(filename)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for column in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
                ws.column_dimensions[get_column_letter(column[0].column)].width = max_length + 2
        wb.save(filename)

        print(f"\n✅ Resultados guardados en {filename}.")

    def generar_individual_graphs_from_excel(self, filename="../results/resultados_simulacion.xlsx"):
        # TODO: revisar los datos que estamos guardando en el excel y ver qué gráficos tiene sentido hacer para ayudarnos
        #       en el análisis
        return
        # """
        # Genera gráficos individuales basados en las métricas de la simulación, incluyendo gráficos adicionales para análisis más detallados.
        # """
        # import matplotlib.pyplot as plt
        # import pandas as pd
        # import os
        #
        # os.makedirs("../results", exist_ok=True)
        # xls = pd.ExcelFile(filename)
        #
        # for sheet_name in xls.sheet_names:
        #     df = pd.read_excel(xls, sheet_name=sheet_name)
        #
        #     # 📊 Gráfico 1: Duración del Episodio vs Episodio
        #     plt.figure(figsize=(10, 6))
        #     plt.plot(df["episode"], df["episode_duration"], label="Duración del episodio", marker="o", color="blue")
        #     plt.title(f"Duración del Episodio vs Episodio\nAlgoritmo: {sheet_name}")
        #     plt.xlabel("Episodio")
        #     plt.ylabel("Duración (ms)")
        #     plt.grid()
        #     plt.legend()
        #     plt.savefig(f"../results/Duracion_Episodio-{sheet_name}.png")
        #     plt.close()
        #
        #     # Gráfico 2: Tasa de Paquetes Entregados por Segundo vs Episodio
        #     plt.figure(figsize=(10, 6))
        #     plt.plot(df["episode"], df["tasa_paquetes_por_segundo"], label="Tasa de entrega (pkt/s)", marker="s", color="green")
        #     plt.title(f"Tasa de Entrega vs Episodio\nAlgoritmo: {sheet_name}")
        #     plt.xlabel("Episodio")
        #     plt.ylabel("Paquetes por segundo")
        #     plt.grid()
        #     plt.legend()
        #     plt.savefig(f"../results/Tasa_Entrega-{sheet_name}.png")
        #     plt.close()
        #
        #     # 📊 Gráfico 3: Hops Promedio por Episodio
        #     plt.figure(figsize=(10, 6))
        #     plt.plot(df["episode"], df["hops_promedio"], label="Hops Promedio", marker="^", color="purple")
        #     plt.title(f"Hops Promedio vs Episodio\nAlgoritmo: {sheet_name}")
        #     plt.xlabel("Episodio")
        #     plt.ylabel("Hops Promedio")
        #     plt.grid()
        #     plt.legend()
        #     plt.savefig(f"../results/Hops_Promedio-{sheet_name}.png")
        #     plt.close()
        #
        #     # Gráfico 4: Distribución de Duración del Episodio
        #     plt.figure(figsize=(10, 6))
        #     plt.hist(df["episode_duration"], bins=20, color="orange", edgecolor="black")
        #     plt.title(f"Distribución de la Duración del Episodio\nAlgoritmo: {sheet_name}")
        #     plt.xlabel("Duración (ms)")
        #     plt.ylabel("Frecuencia")
        #     plt.grid()
        #     plt.savefig(f"../results/Distribucion_Duracion-{sheet_name}.png")
        #     plt.close()
        #
        #     # Gráfico 5: Relación Tasa de Entrega vs Hops Promedio
        #     plt.figure(figsize=(10, 6))
        #     plt.scatter(df["tasa_paquetes_por_segundo"], df["hops_promedio"], label="Relación Tasa vs Hops", color="red")
        #     plt.title(f"Relación Tasa de Entrega vs Hops Promedio\nAlgoritmo: {sheet_name}")
        #     plt.xlabel("Tasa de Entrega (pkt/s)")
        #     plt.ylabel("Hops Promedio")
        #     plt.grid()
        #     plt.legend()
        #     plt.savefig(f"../results/Relacion_Tasa_Hops-{sheet_name}.png")
        #     plt.close()
        #
        #     # Gráfico 6: Actividad de Nodos (Proporción de Nodos Activos)
        #     if "dynamic_changes" in df.columns:
        #         active_nodes = [len(change) if isinstance(change, list) else 0 for change in df["dynamic_changes"]]
        #         plt.figure(figsize=(10, 6))
        #         plt.bar(df["episode"], active_nodes, color="cyan", label="Cambios Dinámicos")
        #         plt.title(f"Actividad de Nodos vs Episodio\nAlgoritmo: {sheet_name}")
        #         plt.xlabel("Episodio")
        #         plt.ylabel("Número de Cambios")
        #         plt.grid()
        #         plt.legend()
        #         plt.savefig(f"../results/Actividad_Nodos-{sheet_name}.png")
        #         plt.close()
        #
        #     # Gráfico 7: Métricas vs Tiempo Real
        #     if "start_time" in df.columns and "end_time" in df.columns:
        #         tiempos_reales = df["end_time"] - df["start_time"]
        #         plt.figure(figsize=(10, 6))
        #         plt.plot(df["episode"], tiempos_reales, label="Duración Real del Episodio", marker="*", color="magenta")
        #         plt.title(f"Duración Real del Episodio vs Episodio\nAlgoritmo: {sheet_name}")
        #         plt.xlabel("Episodio")
        #         plt.ylabel("Duración Real (ms)")
        #         plt.grid()
        #         plt.legend()
        #         plt.savefig(f"../results/Duracion_Real_Episodio-{sheet_name}.png")
        #         plt.close()
        #
        #     # Gráfico 8: Relación entre Episodio y Cambios Dinámicos
        #     plt.figure(figsize=(10, 6))
        #     num_changes = [len(eval(change)) for change in df["dynamic_changes"]]
        #     plt.plot(df["episode"], num_changes, label="Cambios Dinámicos", marker="x", color="orange")
        #     plt.title(f"Cambios Dinámicos vs Episodio\nAlgoritmo: {sheet_name}")
        #     plt.xlabel("Episodio")
        #     plt.ylabel("Número de Cambios")
        #     plt.grid()
        #     plt.legend()
        #     plt.savefig(f"../results/Cambios_Dinamicos-{sheet_name}.png")
        #     plt.close()
        #
        # # Gráfico 9: Comparación de Algoritmos (Promedios Globales)
        # resumen_global = []
        # for sheet_name in xls.sheet_names:
        #     df = pd.read_excel(xls, sheet_name=sheet_name)
        #     promedio_duracion = df["episode_duration"].mean()
        #     promedio_tasa = df["tasa_paquetes_por_segundo"].mean()
        #     promedio_hops = df["hops_promedio"].mean()
        #     resumen_global.append([sheet_name, promedio_duracion, promedio_tasa, promedio_hops])
        #
        # resumen_df = pd.DataFrame(resumen_global, columns=["Algoritmo", "Duración Promedio (ms)", "Tasa Promedio (pkt/s)", "Hops Promedio"])
        #
        # plt.figure(figsize=(10, 6))
        # x = range(len(resumen_df))
        # plt.bar(x, resumen_df["Duración Promedio (ms)"], width=0.3, label="Duración Promedio", align="center")
        # plt.bar(x, resumen_df["Tasa Promedio (pkt/s)"], width=0.3, label="Tasa Promedio", align="edge")
        # plt.bar(x, resumen_df["Hops Promedio"], width=0.3, label="Hops Promedio", align="edge")
        # plt.xticks(x, resumen_df["Algoritmo"], rotation=45)
        # plt.title("Comparación de Métricas Promedio entre Algoritmos")
        # plt.xlabel("Algoritmo")
        # plt.ylabel("Métrica")
        # plt.legend()
        # plt.tight_layout()
        # plt.savefig("../results/Comparacion_Algoritmos.png")
        # plt.close()
        #
        # print("\nGráficos generados en '../results/'.")

    def generar_comparative_graphs_from_excel(self, filename="../results/resultados_simulacion.xlsx"):
        """
        Genera gráficos comparativos basados en las métricas de la simulación, comparando todos los algoritmos en un solo gráfico por métrica.
        """
        import matplotlib.pyplot as plt
        import pandas as pd
        import os

        os.makedirs("../results", exist_ok=True)
        xls = pd.ExcelFile(filename)

        # Diccionario para almacenar datos de todas las hojas
        all_data = {
            "episode_duration": {},
            # "tasa_paquetes_por_segundo": {},
            "hops_promedio": {},
            # "delivered_packets": {},
            # "total_packets": {}
            "total_hops": {},
            "average_delivery_time": {},
            "success_rate": {},
            "episode_success": {}
        }

        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            all_data["episode_duration"][sheet_name] = df["episode_duration"]
            # all_data["tasa_paquetes_por_segundo"][sheet_name] = df["tasa_paquetes_por_segundo"]
            all_data["hops_promedio"][sheet_name] = df["total_hops"] / df["episode"]
            # all_data["delivered_packets"][sheet_name] = df["delivered_packets"]
            # all_data["total_packets"][sheet_name] = df["total_packets"]
            all_data["total_hops"][sheet_name] = df["total_hops"]
            all_data["average_delivery_time"][sheet_name] = df["episode_duration"] / df["total_hops"]
            all_data["success_rate"][sheet_name] = df["episode_success"].fillna(False).infer_objects(copy=False)
            all_data["episode_success"][sheet_name] = df["episode_success"]
            
        # Gráfico comparativo de Duración del Episodio
        plt.figure(figsize=(10, 6))
        for algorithm, data in all_data["episode_duration"].items():
            plt.plot(data, label=f"{algorithm}")
        plt.title("Comparación de Duración del Episodio entre Algoritmos")
        plt.xlabel("Episodio")
        plt.ylabel("Duración (ms)")
        plt.grid()
        plt.legend()
        plt.savefig("../results/Comparacion_Duracion_Episodio.png")
        plt.close()

        # Gráfico comparativo de Tasa de Paquetes Entregados por Segundo
        # plt.figure(figsize=(10, 6))
        # for algorithm, data in all_data["tasa_paquetes_por_segundo"].items():
        #     plt.plot(data, label=f"{algorithm}")
        # plt.title("Comparación de Tasa de Entrega entre Algoritmos")
        # plt.xlabel("Episodio")
        # plt.ylabel("Paquetes por segundo")
        # plt.grid()
        # plt.legend()
        # plt.savefig("../results/Comparacion_Tasa_Entrega.png")
        # plt.close()

        # Gráfico comparativo de Hops Promedio
        plt.figure(figsize=(10, 6))
        for algorithm, data in all_data["hops_promedio"].items():
            plt.plot(data, label=f"{algorithm}")
        plt.title("Comparación de Hops Promedio entre Algoritmos")
        plt.xlabel("Episodio")
        plt.ylabel("Hops Promedio")
        plt.grid()
        plt.legend()
        plt.savefig("../results/Comparacion_Hops_Promedio.png")
        plt.close()

        # Gráfico comparativo de Paquetes Entregados y Totales
        # algorithms = list(all_data["delivered_packets"].keys())
        # delivered_packets = [all_data["delivered_packets"][alg].sum() for alg in algorithms]
        # total_packets = [all_data["total_packets"][alg].sum() for alg in algorithms]

        # x = np.arange(len(algorithms))  # the label locations
        # width = 0.35  # the width of the bars

        # fig, ax = plt.subplots(figsize=(10, 6))
        # rects1 = ax.bar(x - width/2, delivered_packets, width, label='Paquetes Entregados')
        # rects2 = ax.bar(x + width/2, total_packets, width, label='Total de Paquetes')

        # # Add some text for labels, title and custom x-axis tick labels, etc.
        # ax.set_xlabel('Algoritmo')
        # ax.set_ylabel('Número de Paquetes')
        # ax.set_title('Comparación de Paquetes Entregados y Totales entre Algoritmos')
        # ax.set_xticks(x)
        # ax.set_xticklabels(algorithms)
        # ax.legend()

        # fig.tight_layout()
        # plt.grid()
        # plt.savefig("./results/Comparacion_Paquetes_Entregados_Totales.png")
        # plt.close()

        # Gráfico comparativo de Tiempo Promedio de Entrega
        plt.figure(figsize=(10, 6))
        for algorithm, data in all_data["average_delivery_time"].items():
            plt.plot(data, label=f"{algorithm}")
        plt.title("Comparación de Tiempo Promedio de Entrega entre Algoritmos")
        plt.xlabel("Episodio")
        plt.ylabel("Tiempo Promedio de Entrega (ms)")
        plt.grid()
        plt.legend()
        plt.savefig("../results/Comparacion_Tiempo_Promedio_Entrega.png")
        plt.close()

        # Gráfico comparativo de Tasa de Éxito
        plt.figure(figsize=(10, 6))
        for algorithm, data in all_data["success_rate"].items():
            plt.plot(data, label=f"{algorithm}")
        plt.title("Comparación de Tasa de Éxito entre Algoritmos")
        plt.xlabel("Episodio")
        plt.ylabel("Tasa de Éxito")
        plt.grid()
        plt.legend()
        plt.savefig("../results/Comparacion_Tasa_Exito.png")
        plt.close()

        # Gráfico de columnas verticales de Tasa de Éxito por Episodio
        plt.figure(figsize=(10, 6))

        # Initialize dictionaries to count TRUE, FALSE, and blank values for each algorithm
        success_counts = {algorithm: {"TRUE": 0, "FALSE": 0, "BLANK": 0} for algorithm in all_data["episode_success"].keys()}

        # Count the occurrences of TRUE, FALSE, and blank values for each algorithm
        for algorithm, data in all_data["episode_success"].items():
            for value in data:
                if pd.isna(value):
                    success_counts[algorithm]["BLANK"] += 1
                elif value:
                    success_counts[algorithm]["TRUE"] += 1
                else:
                    success_counts[algorithm]["FALSE"] += 1

        # Plot the counts for each algorithm
        bar_width = 0.25
        index = np.arange(len(success_counts))

        # Plot TRUE values
        plt.bar(index, [success_counts[alg]["TRUE"] for alg in success_counts], bar_width, label="TRUE")

        # Plot FALSE values
        plt.bar(index + bar_width, [success_counts[alg]["FALSE"] for alg in success_counts], bar_width, label="FALSE")

        # Plot Blank values
        plt.bar(index + 2 * bar_width, [success_counts[alg]["BLANK"] for alg in success_counts], bar_width, label="BLANK")

        plt.title("Tasa de Éxito por Episodio entre Algoritmos")
        plt.xlabel("Algoritmo")
        plt.ylabel("Cantidad")
        plt.xticks(index + bar_width, success_counts.keys())
        plt.grid()
        plt.legend()
        plt.savefig("../results/Tasa_Exito_Columnas.png")
        plt.close()

        print("\nGráficos comparativos generados en '../results/'.")


    def generate_heat_map(self, q_tables):
        q_table_data = []
        for q_table in q_tables:
            for state, actions in q_table.items():
                for action, q_value in actions.items():
                    q_table_data.append((state, action, q_value))

        if not q_table_data:
            print(f"No Q-table data available.")
            return

        # Extract unique states and actions
        states = sorted(set(state for state, _, _ in q_table_data))
        actions = sorted(set(action for _, action, _ in q_table_data))

        # Create a matrix to hold Q-values
        q_matrix = np.zeros((len(states), len(actions)))

        for state, action, q_value in q_table_data:
            state_index = states.index(state)
            action_index = actions.index(action)
            q_matrix[state_index, action_index] = q_value

        fig, ax = plt.subplots()
        cax = ax.imshow(q_matrix, cmap='RdYlGn', aspect='auto', vmin=0, vmax=1)

        # Add color bar
        fig.colorbar(cax)

        # Set axis labels
        ax.set_xticks(np.arange(len(actions)))
        ax.set_yticks(np.arange(len(states)))
        ax.set_xticklabels(actions)
        ax.set_yticklabels(states)

        # Annotate each cell with its value
        for i in range(len(states)):
            for j in range(len(actions)):
                ax.text(j, i, f'{q_matrix[i, j]:.2f}', ha='center', va='center', color='black')

        plt.xlabel('Actions (Next Node ID)')
        plt.ylabel('States (Node ID)')
        plt.title(f'Q-Table Heat Map')

        output_folder = '../results'
        # Save the heat map as a .png file
        filename = os.path.join(output_folder, f'q_table_heat_map.png')
        plt.savefig(filename)
        plt.close()

        print(f'Heat map saved to {filename}')