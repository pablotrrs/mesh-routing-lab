import datetime
import json
import logging as log
import os
from typing import Dict, List, Optional, Union

import matplotlib.pyplot as plt
import numpy as np
from core.packet_registry import registry


class MetricsManager:
    """Manages and stores metrics for the simulation.

    Attributes:
        metrics (Dict[str, Union[int, float, str, List, Dict]]): Dictionary to store simulation metrics.
    """

    def __init__(self) -> None:
        """Initializes the MetricsManager with an empty metrics dictionary."""
        self.metrics: Dict[str, Union[int, float, str, List, Dict]] = {}
        log.debug("MetricsManager initialized.")

    def initialize(
        self,
        max_hops: int,
        topology_file: str,
        functions_sequence: List[str],
        mean_disconnection_interval_ms: float,
        mean_reconnection_interval_ms: float,
        disconnection_probability: float,
        algorithms: List[str],
        penalty: float,
    ) -> None:
        """Initializes the metrics for a new simulation with multiple algorithms.

        Args:
            max_hops (int): Maximum number of hops allowed.
            topology_file (str): Path to the topology file.
            functions_sequence (List[str]): Sequence of node functions.
            mean_disconnection_interval_ms (float): Mean interval for dynamic changes.
            mean_reconnection_interval_ms (float): Interval for node reconnection.
            disconnection_probability (float): Probability of node disconnection.
            algorithms (List[str]): List of algorithms used in the simulation.
            penalty (float): Penalty for Q-Routing.
        """
        self.metrics = {
            "simulation_id": 1,
            "parameters": {
                "max_hops": max_hops,
                "algorithms": algorithms,
                "mean_disconnection_interval_ms": mean_disconnection_interval_ms,
                "mean_reconnection_interval_ms": mean_reconnection_interval_ms,
                "topology_file": topology_file,
                "functions_sequence": [func.value for func in functions_sequence],
                "disconnection_probability": disconnection_probability,
            },
            "total_time": None,
            "runned_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

        for algorithm in algorithms:
            self.metrics[algorithm] = {"success_rate": 0.0, "episodes": []}

            if algorithm == "Q_ROUTING":
                self.metrics[algorithm]["penalty"] = penalty

        log.debug("Metrics initialized for simulation.")

    def log_episode(
        self,
        algorithm: str,
        episode_number: int,
        start_time: float,
        end_time: float,
        episode_success: bool,
        route: List[str],
        total_hops: int,
        dynamic_changes: List[Dict],
    ) -> None:
        """Logs the metrics of an episode for a specific algorithm.

        Args:
            algorithm (str): The algorithm used in the episode.
            episode_number (int): The number of the episode.
            start_time (float): Start time of the episode.
            end_time (float): End time of the episode.
            episode_success (bool): Whether the episode was successful.
            route (List[str]): The route taken in the episode.
            total_hops (int): Total number of hops in the episode.
            dynamic_changes (List[Dict]): List of dynamic changes during the episode.
        """
        if algorithm not in self.metrics:
            self.metrics[algorithm] = {"success_rate": 0.0, "episodes": []}

        self.metrics[algorithm]["episodes"].append(
            {
                "episode_number": episode_number,
                "start_time": start_time,
                "end_time": end_time,
                "episode_duration": end_time - start_time,
                "episode_success": episode_success,
                "route": route,
                "total_hops": total_hops,
                "dynamic_changes": dynamic_changes,
                "dynamic_changes_count": len(dynamic_changes),
            }
        )

        log.debug(f"Logged episode {episode_number} for algorithm {algorithm}.")

    def finalize_simulation(
        self, total_time: float, successful_episodes: int, episodes: int
    ) -> None:
        """Finalizes the simulation and saves the results.

        Args:
            total_time (float): Total time taken for the simulation.
            successful_episodes (int): Number of successful episodes.
            episodes (int): Total number of episodes.
        """
        algorithm = self.metrics["parameters"]["algorithms"][-1]
        self.metrics[algorithm]["success_rate"] = (
            successful_episodes / episodes if episodes > 0 else 0.0
        )
        self.metrics["total_time"] = total_time

        self.save_metrics_to_file()
        self.save_results_to_excel()
        log.debug("Simulation finalized and results saved.")

    def save_metrics_to_file(self, directory: str = "../resources/results/single-run") -> None:
        """Saves the simulation metrics to a JSON file.

        Args:
            directory (str): Directory to save the file. Defaults to "../resources/results/single-run".
        """
        os.makedirs(directory, exist_ok=True)

        filename = f"{directory}/simulation_{self.metrics['simulation_id']}.json"

        with open(filename, "w", encoding="utf-8") as file:
            json.dump(self.metrics, file, indent=4)

        log.debug(f"Simulation metrics saved to {filename}.")

    def save_results_to_excel(
        self, filename: str = "../resources/results/resultados_simulacion.xlsx"
    ) -> None:
        """Saves the simulation results to an Excel file.

        Args:
            filename (str): Path to the Excel file. Defaults to "../../resources/results/resultados_simulacion.xlsx".
        """
        import json
        import os

        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        os.makedirs("../../resources/results", exist_ok=True)

        if os.path.exists(filename):
            try:
                pd.ExcelFile(filename)
            except Exception:
                log.error(
                    f"Corrupted file detected: {filename}. Deleting and regenerating..."
                )
                os.remove(filename)

        if not self.metrics:
            log.error("self.metrics is empty")
            return

        metrics_data = {
            algorithm: {
                "episode": [],
                "start_time": [],
                "end_time": [],
                "episode_duration": [],
                "episode_success": [],
                "total_hops": [],
                "dynamic_changes": [],
                "packet_log_raw": [],
            }
            for algorithm in self.metrics.keys()
            if algorithm
            not in ["simulation_id", "parameters", "total_time", "runned_at"]
        }
        log.debug(
            f"Algorithms found in metrics: {list(metrics_data.keys())}"
        )

        for algorithm, episodes in self.metrics.items():
            if algorithm in ["simulation_id", "parameters", "total_time", "runned_at"]:
                continue
            for episode_data in episodes["episodes"]:
                log.debug(
                    f"Processing episode no {episode_data['episode_number']} for algorithm {algorithm}..."
                )
                episode_number = episode_data["episode_number"]

                packet_log = registry.packet_log.get(episode_number, {})

                metrics_data[algorithm]["episode"].append(episode_number)
                metrics_data[algorithm]["start_time"].append(
                    episode_data.get("start_time", 0)
                )
                metrics_data[algorithm]["end_time"].append(
                    episode_data.get("end_time", 0)
                )
                metrics_data[algorithm]["episode_duration"].append(
                    episode_data.get("episode_duration", 0)
                )
                metrics_data[algorithm]["episode_success"].append(
                    episode_data.get("episode_success", False)
                )
                metrics_data[algorithm]["total_hops"].append(
                    episode_data.get("total_hops", 0)
                )
                metrics_data[algorithm]["dynamic_changes"].append(
                    len(episode_data.get("dynamic_changes", []))
                )

                try:
                    packet_log_json = json.dumps(packet_log, indent=2)
                except Exception as e:
                    log.error(e)
                    packet_log_json = "{}"

                metrics_data[algorithm]["packet_log_raw"].append(packet_log_json)

        with pd.ExcelWriter(filename, engine="openpyxl", mode="w") as writer:
            for algorithm, data in metrics_data.items():
                if not data["episode"]:
                    log.error(f"no episodes for algorithm {algorithm}.")
                    continue

                df = pd.DataFrame(data)
                df.to_excel(writer, sheet_name=algorithm, index=False)

        wb = load_workbook(filename)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for column in ws.columns:
                max_length = max(
                    len(str(cell.value)) if cell.value else 0 for cell in column
                )
                ws.column_dimensions[get_column_letter(column[0].column)].width = (
                    max_length + 2
                )
        wb.save(filename)

        log.debug(f"\nresults saved in {filename}.")

    def generar_comparative_graphs_from_excel(
        self, filename: str = "../resources/results/resultados_simulacion.xlsx"
    ) -> None:
        """Generates comparative graphs based on simulation metrics.

        Args:
            filename (str): Path to the Excel file. Defaults to "../resources/results/resultados_simulacion.xlsx".
        """
        import os

        import matplotlib.pyplot as plt
        import pandas as pd

        os.makedirs("../resources/results", exist_ok=True)
        xls = pd.ExcelFile(filename)

        all_data = {
            "episode_duration": {},
            "hops_promedio": {},
            "total_hops": {},
            "average_delivery_time": {},
            "success_rate": {},
            "episode_success": {},
        }

        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            all_data["episode_duration"][sheet_name] = df["episode_duration"]
            all_data["hops_promedio"][sheet_name] = df["total_hops"] / df["episode"]
            all_data["total_hops"][sheet_name] = df["total_hops"]
            all_data["average_delivery_time"][sheet_name] = (
                df["episode_duration"] / df["total_hops"]
            )
            all_data["success_rate"][sheet_name] = (
                df["episode_success"].fillna(False).infer_objects(copy=False)
            )
            all_data["episode_success"][sheet_name] = df["episode_success"]

        plt.figure(figsize=(10, 6))
        for algorithm, data in all_data["episode_duration"].items():
            plt.plot(data, label=f"{algorithm}")
        plt.title("Comparación de Duración del Episodio entre Algoritmos")
        plt.xlabel("Episodio")
        plt.ylabel("Duración (ms)")
        plt.grid()
        plt.legend()
        plt.savefig("../resources/results/Comparacion_Duracion_Episodio.png")
        plt.close()

        plt.figure(figsize=(10, 6))
        for algorithm, data in all_data["hops_promedio"].items():
            plt.plot(data, label=f"{algorithm}")
        plt.title("Comparación de Hops Promedio entre Algoritmos")
        plt.xlabel("Episodio")
        plt.ylabel("Hops Promedio")
        plt.grid()
        plt.legend()
        plt.savefig("../resources/results/Comparacion_Hops_Promedio.png")
        plt.close()

        plt.figure(figsize=(10, 6))
        for algorithm, data in all_data["average_delivery_time"].items():
            plt.plot(data, label=f"{algorithm}")
        plt.title("Comparación de Tiempo Promedio de Entrega entre Algoritmos")
        plt.xlabel("Episodio")
        plt.ylabel("Tiempo Promedio de Entrega (ms)")
        plt.grid()
        plt.legend()
        plt.savefig("../resources/results/Comparacion_Tiempo_Promedio_Entrega.png")
        plt.close()

        plt.figure(figsize=(10, 6))
        for algorithm, data in all_data["success_rate"].items():
            plt.plot(data, label=f"{algorithm}")
        plt.title("Comparación de Tasa de Éxito entre Algoritmos")
        plt.xlabel("Episodio")
        plt.ylabel("Tasa de Éxito")
        plt.grid()
        plt.legend()
        plt.savefig("../resources/results/Comparacion_Tasa_Exito.png")
        plt.close()

        plt.figure(figsize=(10, 6))

        success_counts = {
            algorithm: {"TRUE": 0, "FALSE": 0, "BLANK": 0}
            for algorithm in all_data["episode_success"].keys()
        }

        for algorithm, data in all_data["episode_success"].items():
            for value in data:
                if pd.isna(value):
                    success_counts[algorithm]["BLANK"] += 1
                elif value:
                    success_counts[algorithm]["TRUE"] += 1
                else:
                    success_counts[algorithm]["FALSE"] += 1

        bar_width = 0.25
        index = np.arange(len(success_counts))

        plt.bar(
            index,
            [success_counts[alg]["TRUE"] for alg in success_counts],
            bar_width,
            label="TRUE",
        )

        plt.bar(
            index + bar_width,
            [success_counts[alg]["FALSE"] for alg in success_counts],
            bar_width,
            label="FALSE",
        )

        plt.bar(
            index + 2 * bar_width,
            [success_counts[alg]["BLANK"] for alg in success_counts],
            bar_width,
            label="BLANK",
        )

        plt.title("Tasa de Éxito por Episodio entre Algoritmos")
        plt.xlabel("Algoritmo")
        plt.ylabel("Cantidad")
        plt.xticks(index + bar_width, success_counts.keys())
        plt.grid()
        plt.legend()
        plt.savefig("../resources/results/Tasa_Exito_Columnas.png")
        plt.close()

        log.debug("Comparative graphs generated in: '../resources/results/'.")

    def generate_heat_map(self, q_tables):
        q_table_data = []
        for q_table in q_tables:
            for state, actions in q_table.items():
                for action, q_value in actions.items():
                    q_table_data.append((state, action, q_value))

        if not q_table_data:
            log.error(f"No Q-table data available.")
            return

        states = sorted(set(state for state, _, _ in q_table_data))
        actions = sorted(set(action for _, action, _ in q_table_data))

        q_matrix = np.zeros((len(states), len(actions)))

        for state, action, q_value in q_table_data:
            state_index = states.index(state)
            action_index = actions.index(action)
            q_matrix[state_index, action_index] = q_value

        fig, ax = plt.subplots()
        cax = ax.imshow(q_matrix, cmap="RdYlGn", aspect="auto", vmin=0, vmax=1)

        fig.colorbar(cax)

        ax.set_xticks(np.arange(len(actions)))
        ax.set_yticks(np.arange(len(states)))
        ax.set_xticklabels(actions)
        ax.set_yticklabels(states)

        for i in range(len(states)):
            for j in range(len(actions)):
                ax.text(
                    j,
                    i,
                    f"{q_matrix[i, j]:.2f}",
                    ha="center",
                    va="center",
                    color="black",
                )

        plt.xlabel("Actions (Next Node ID)")
        plt.ylabel("States (Node ID)")
        plt.title(f"Q-Table Heat Map")

        output_folder = "../resources/results"
        filename = os.path.join(output_folder, f"q_table_heat_map.png")
        plt.savefig(filename)
        plt.close()

        log.debug(f"Heat map saved to {filename}")
