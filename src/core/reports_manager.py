import json
import logging as log
import os
from typing import Dict, List, Union

import matplotlib.pyplot as plt
import numpy as np

from core.base import Algorithm, SimulationConfig


class ReportsManager:
    """Manages and stores metrics for the simulation.

    Attributes:
        metrics (Dict[str, Union[int, float, str, List, Dict]]): Dictionary to store simulation metrics.
    """

    def __init__(self) -> None:
        """Initializes the MetricsManager with an empty metrics dictionary."""
        self.metrics: Dict[str, Union[int, float, str, List, Dict]] = {}
        self.config: SimulationConfig = None
        self.results_dir = self.get_next_results_directory()
        log.debug("MetricsManager initialized.")

    def generate_reports(self):
        filename = self._save_metrics_to_file(self.results_dir)
        self._save_results_to_excel(os.path.join(self.results_dir, "resultados_simulacion.xlsx"))
        self._generate_comparative_graphs_from_excel(os.path.join(self.results_dir, "resultados_simulacion.xlsx"))
        self.generate_q_table_heatmap(self.results_dir)

        if Algorithm.Q_ROUTING in self.config.algorithms:
           self.generate_q_routing_policy_graphs_from_metrics(
               filename,
               os.path.join(self.results_dir, "Algorithm.Q_ROUTING")
           )

    @staticmethod
    def get_next_results_directory(base_path="../resources/results"):
        os.makedirs(base_path, exist_ok=True)
        index = 1
        while True:
            candidate = os.path.join(base_path, str(index))
            if not os.path.exists(candidate):
                os.makedirs(candidate)
                return candidate
            index += 1

    def _save_metrics_to_file(
        self, directory: str = "../resources/results/single-run"
    ) -> str:
        """Saves the simulation metrics to a JSON file.

        Args:
            directory (str): Directory to save the file. Defaults to "../resources/results/single-run".
        """
        os.makedirs(directory, exist_ok=True)

        from core.packet_registry import registry

        filename = f"{directory}/metrics.json"

        with open(filename, "w", encoding="utf-8") as file:
            json.dump(registry.metrics, file, indent=4)

        log.debug(f"Simulation metrics saved to {filename}.")
        return filename

    def _save_results_to_excel(
        self, filename: str = "../resources/results/resultados_simulacion.xlsx"
    ) -> None:
        """Saves the simulation results to an Excel file.

        Args:
            filename (str): Path to the Excel file. Defaults to "../../resources/results/resultados_simulacion.xlsx".
        """
        import json
        import os

        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        os.makedirs("../../resources/results", exist_ok=True)

        if os.path.exists(filename):
            try:
                pd.ExcelFile(filename)
            except Exception:
                log.error(
                    f"Corrupted file detected: {filename}. Deleting and regenerating..."
                )
                os.remove(filename)

        from core.packet_registry import registry

        metrics = registry.metrics
        if not metrics:
            log.error("metrics are empty")
            return

        metrics_data = {
            algorithm: {
                "episode": [],
                "start_time": [],
                "end_time": [],
                "episode_duration": [],
                "episode_success": [],
                "total_hops": [],
                "dynamic_changes": [],
                "packet_log_raw": [],
            }
            for algorithm in metrics.keys()
            if algorithm
            not in ["simulation_id", "parameters", "total_time", "runned_at"]
        }
        log.debug(f"Algorithms found in metrics: {list(metrics_data.keys())}")

        for algorithm, episodes in metrics.items():
            if algorithm in ["simulation_id", "parameters", "total_time", "runned_at"]:
                continue
            for episode_data in episodes["episodes"]:
                log.debug(
                    f"Processing episode no {episode_data['episode_number']} for algorithm {algorithm}..."
                )
                episode_number = episode_data["episode_number"]

                packet_log = registry.packet_log.get(episode_number, {})

                metrics_data[algorithm]["episode"].append(episode_number)
                metrics_data[algorithm]["start_time"].append(
                    episode_data.get("start_time", 0)
                )
                metrics_data[algorithm]["end_time"].append(
                    episode_data.get("end_time", 0)
                )
                metrics_data[algorithm]["episode_duration"].append(
                    episode_data.get("episode_duration", 0)
                )
                metrics_data[algorithm]["episode_success"].append(
                    episode_data.get("episode_success", False)
                )
                metrics_data[algorithm]["total_hops"].append(
                    episode_data.get("total_hops", 0)
                )
                metrics_data[algorithm]["dynamic_changes"].append(
                    len(episode_data.get("dynamic_changes", []))
                )

                try:
                    packet_log_json = json.dumps(packet_log, indent=2)
                except Exception as e:
                    log.error(e)
                    packet_log_json = "{}"

                metrics_data[algorithm]["packet_log_raw"].append(packet_log_json)

        with pd.ExcelWriter(filename, engine="openpyxl", mode="w") as writer:
            for algorithm, data in metrics_data.items():
                if not data["episode"]:
                    log.debug(f"no episodes for algorithm {algorithm}.")
                    continue

                df = pd.DataFrame(data)
                df.to_excel(writer, sheet_name=str(algorithm), index=False)

        wb = load_workbook(filename)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for column in ws.columns:
                max_length = max(
                    len(str(cell.value)) if cell.value else 0 for cell in column
                )
                ws.column_dimensions[get_column_letter(column[0].column)].width = (
                    max_length + 2
                )
        wb.save(filename)

        log.debug(f"\nresults saved in {filename}.")

    def _generate_comparative_graphs_from_excel(self, filename: str = "../resources/results/resultados_simulacion.xlsx") -> None:
        import os
        import matplotlib.pyplot as plt
        import pandas as pd
        import numpy as np

        def create_config_label(config):
            import os
            from core.enums import Algorithm

            label = (
                f"Episodios: {config.episodes}\n"
                f"Max hops: {config.max_hops}\n"
                f"Timeout episodio: {config.episode_timeout_ms} ms\n"
                f"Prob. desconexión: {config.disconnection_probability}\n"
                f"Int. desconexión fija: {config.disconnection_interval_ms} ms\n"
                f"Int. reconexión fija: {config.reconnection_interval_ms} ms\n"
                f"Int. desconexión media: {config.mean_disconnection_interval_ms} ms\n"
                f"Int. reconexión media: {config.mean_reconnection_interval_ms} ms\n"
                f"Topología: {os.path.basename(config.topology_file)}\n"
                f"Secuencia de funciones: {' → '.join(f.value for f in config.functions_sequence)}"
            )

            if Algorithm.Q_ROUTING in config.algorithms:
                label += "\n"
                label += f"Usa epsilon decay: {'Sí' if config.use_epsilon_decay else 'No'}\n"
                if config.use_epsilon_decay:
                    label += (
                        f"Epsilon inicial: {config.initial_epsilon:.2f}\n"
                        f"Factor de decay: {config.epsilon_decay:.4f}\n"
                        f"Epsilon mínimo: {config.epsilon_min:.2f}"
                    )
                else:
                    label += f"Epsilon fijo: {config.fixed_epsilon:.2f}"

            return label

        os.makedirs(self.results_dir, exist_ok=True)
        xls = pd.ExcelFile(filename)
        all_data = {
            "episode_duration": {},
            "hops_promedio": {},
            "total_hops": {},
            "average_delivery_time": {},
            "success_rate": {},
            "episode_success": {},
        }

        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            all_data["episode_duration"][sheet_name] = df["episode_duration"]
            all_data["hops_promedio"][sheet_name] = df["total_hops"] / df["episode"]
            all_data["total_hops"][sheet_name] = df["total_hops"]
            all_data["average_delivery_time"][sheet_name] = df["episode_duration"] / df["total_hops"]
            all_data["success_rate"][sheet_name] = df["episode_success"].fillna(False)
            all_data["episode_success"][sheet_name] = df["episode_success"]

        algorithm_names = list(all_data["episode_duration"].keys())
        output_dirs = {name: os.path.join(self.results_dir, name) for name in algorithm_names}
        output_dirs["all"] = os.path.join(self.results_dir, "all")
        for path in output_dirs.values():
            os.makedirs(path, exist_ok=True)

        def save_line_chart(data_dict, title, ylabel, filename, target_dirs):
            label = create_config_label(self.config)
            plt.figure(figsize=(16, 8), dpi=150)
            for algorithm, data in data_dict.items():
                plt.plot(data, label=algorithm, linewidth=2, alpha=0.8)
            plt.title(title)
            plt.xlabel("Episodio")
            plt.ylabel(ylabel)
            plt.grid(True)
            plt.legend()
            plt.annotate(label, xy=(1.01, 0), xycoords='axes fraction', fontsize=10,
                        ha='left', va='bottom', linespacing=1.5,
                        bbox=dict(facecolor='white', alpha=0.7, edgecolor='gray'))
            plt.tight_layout()
            plt.savefig(os.path.join(target_dirs["all"], filename))
            plt.close()

            for algorithm, data in data_dict.items():
                plt.figure(figsize=(16, 8), dpi=150)
                plt.plot(data, label=algorithm, linewidth=2, alpha=0.8)
                plt.title(f"{title} - {algorithm}")
                plt.xlabel("Episodio")
                plt.ylabel(ylabel)
                plt.grid(True)
                plt.legend()
                plt.annotate(label, xy=(1.01, 0), xycoords='axes fraction', fontsize=10,
                            ha='left', va='bottom', linespacing=1.5,
                            bbox=dict(facecolor='white', alpha=0.7, edgecolor='gray'))
                plt.tight_layout()
                plt.savefig(os.path.join(target_dirs[algorithm], filename))
                plt.close()

        save_line_chart(all_data["episode_duration"], "Duración del Episodio", "Duración (ms)", "Duracion_Episodio.png", output_dirs)
        save_line_chart(all_data["total_hops"], "Cantidad de Hops por Episodio", "Cantidad de Hops", "Total_Hops_Episodio.png", output_dirs)

        # Tasa de éxito por algoritmo
        plt.figure(figsize=(12, 8), dpi=150)
        success_counts = {alg: {"TRUE": 0, "FALSE": 0} for alg in all_data["episode_success"]}
        for alg, data in all_data["episode_success"].items():
            for val in data:
                success_counts[alg]["TRUE" if val else "FALSE"] += 1

        bar_width = 0.25
        index = np.arange(len(success_counts))
        true_values = [success_counts[alg]["TRUE"] for alg in success_counts]
        false_values = [success_counts[alg]["FALSE"] for alg in success_counts]
        bars_true = plt.bar(index, true_values, bar_width, label="TRUE")
        bars_false = plt.bar(index + bar_width, false_values, bar_width, label="FALSE")

        for bar in bars_true + bars_false:
            plt.text(bar.get_x() + bar.get_width() / 2, bar.get_height(), f"{int(bar.get_height())}", ha="center", va="bottom", fontsize=8)

        plt.title("Tasa de Éxito por Algoritmo")
        plt.xlabel("Algoritmo")
        plt.ylabel("Cantidad")
        plt.xticks(index + bar_width / 2, success_counts.keys())
        plt.grid(True)
        plt.legend()
        plt.annotate(create_config_label(self.config), xy=(1.01, 0), xycoords='axes fraction', fontsize=10,
                    ha='left', va='bottom', linespacing=1.5,
                    bbox=dict(facecolor='white', alpha=0.7, edgecolor='gray'))
        plt.tight_layout()
        plt.savefig(os.path.join(output_dirs["all"], "Tasa_Exito_Columnas.png"))
        plt.close()

        # Éxitos acumulados
        plt.figure(figsize=(12, 6))
        for alg, successes in all_data["episode_success"].items():
            cumulative = np.cumsum([1 if s else 0 for s in successes])
            plt.plot(cumulative, label=f"{alg}", linewidth=1.5)

        plt.title("Evolución Acumulada de Éxitos")
        plt.xlabel("Episodio")
        plt.ylabel("Éxitos Acumulados")
        plt.grid(True, linestyle='--', alpha=0.5)
        plt.legend()
        plt.annotate(create_config_label(self.config), xy=(1.01, 0), xycoords='axes fraction', fontsize=10,
                    ha='left', va='bottom', linespacing=1.5,
                    bbox=dict(facecolor='white', alpha=0.7, edgecolor='gray'))
        plt.tight_layout()
        plt.savefig(os.path.join(output_dirs["all"], "Evolucion_Success_Acumulado.png"))
        plt.close()

    def generate_q_table_heatmap(
        self,
        directory: str = "../resources/results",
        algorithm="Q_ROUTING"
    ):
        """Generates heatmaps for Q-tables across episodes and creates a GIF to visualize the evolution of Q-values.

        Args:
            directory (str): The directory containing the metrics JSON file. Defaults to "../resources/results".
            algorithm (str): The algorithm name to extract Q-value data from the metrics file. Defaults to "Q_ROUTING".
        Raises:
            ValueError: If the specified algorithm is not found in the metrics file.
        Outputs:
            - Heatmap images for each episode saved in the results directory.
            - A GIF visualizing the evolution of Q-values across episodes.
        Notes:
            - The heatmaps display Q-values between nodes, with NaN values represented as black cells.
            - The function calculates global min, max, and median Q-values for consistent color scaling.
            - Requires the `imageio` library to generate the GIF. If not installed, the GIF generation will be skipped.
        """

        json_file = f"{directory}/metrics.json"

        with open(json_file, "r", encoding="utf-8") as file:
            data = json.load(file)

        if algorithm not in data:
            raise ValueError(f"Algorithm '{algorithm}' not found in the metrics file.")

        episodes = data[algorithm]["episodes"]

        max_node = 0
        q_values = []
        heatmap_paths = []
        for episode in episodes:
            for route in episode["route"]:
                max_node = max(max_node, route["from"], route["to"])
                if "q_value" in route:
                    q_values.append(route["q_value"])
        num_nodes = max_node + 1 

        # Calculate the global min, max, and median Q-values
        min_q_value = min(q_values) if q_values else 0
        max_q_value = max(q_values) if q_values else 1
        median_q_value = np.median(q_values) if q_values else (min_q_value + max_q_value) / 2

        persistent_q_table = np.full((num_nodes, num_nodes), np.nan)

        # Generate a heatmap for each episode
        for episode_index, episode in enumerate(episodes):
            q_table = np.copy(persistent_q_table)

            for route in episode["route"]:
                if "q_value" in route:
                    q_table[route["from"], route["to"]] = route["q_value"]

            persistent_q_table = np.copy(q_table)

            # Generate the heatmap
            plt.figure(figsize=(10, 8))
            masked_q_table = np.ma.masked_where(np.isnan(persistent_q_table), persistent_q_table)
            cmap = plt.cm.RdYlGn
            cmap.set_bad(color='black')  # Set NaN cells to black
            plt.imshow(
                masked_q_table,
                cmap=cmap,
                interpolation="nearest",
                vmin=min_q_value,
                vmax=max_q_value
            )
            plt.colorbar(label="Q-Value")
            plt.title(f"Q-Table Heatmap for {algorithm} - Episode {episode_index + 1}")
            plt.xlabel("To Node")
            plt.ylabel("From Node")
            plt.xticks(range(num_nodes))
            plt.yticks(range(num_nodes))

            # Annotate the heatmap with actual Q-values
            for i in range(num_nodes):
                for j in range(num_nodes):
                    value = persistent_q_table[i, j]
                    if not np.isnan(value):
                        plt.text(j, i, f"{value:.2f}", ha="center", va="center", color="black")

            from core.enums import Algorithm
            algorithm_enum = Algorithm[algorithm]
            output_dir = os.path.join(self.results_dir, str(algorithm_enum), "q-table")
            os.makedirs(output_dir, exist_ok=True)
            heatmap_path = os.path.join(output_dir, f"q_table_heatmap_episode_{episode_index + 1}.png")
            plt.tight_layout()
            plt.savefig(heatmap_path)
            plt.close()
            heatmap_paths.append(heatmap_path)
            log.debug(f"Q-Table heatmap for episode {episode_index + 1} saved to {heatmap_path}")

        # Generate a GIF from the heatmaps
        import imageio
        gif_path = os.path.join(output_dir, f"q_table_heatmap.gif")
        try:
            with imageio.get_writer(gif_path, mode='I', duration=5) as writer:
                for path in heatmap_paths:
                    image = imageio.imread(path)
                    writer.append_data(image)
            log.debug(f"GIF of Q-Table heatmaps saved to {gif_path}")
        except ImportError:
            log.error("imageio library is required to generate GIFs. Please install it using 'pip install imageio'.")

    def generate_q_routing_policy_graphs_from_metrics(self, metrics_path: str, output_dir: str = "./q_routing_policy_analysis") -> None:
        import os
        import json
        import numpy as np
        import matplotlib.pyplot as plt

        os.makedirs(output_dir, exist_ok=True)

        with open(metrics_path, "r") as f:
            metrics = json.load(f)

        q_data = metrics.get("Q_ROUTING", {}).get("episodes", [])
        if not q_data:
            print("No Q_ROUTING data found in metrics.")
            return

        explore_ratios = []
        avg_epsilons = []
        heatmap_data = []
        explore_durations = []
        exploit_durations = []

        for episode in q_data:
            route = episode.get("route", [])
            explore_count = 0
            total = 0
            epsilons = []

            row = []

            for hop in route:
                decision = hop.get("policy_decision")
                epsilon = hop.get("epsilon")
                if decision is not None:
                    row.append(1 if decision == "EXPLORE" else 0)
                    if decision == "EXPLORE":
                        explore_count += 1
                if epsilon is not None:
                    epsilons.append(epsilon)
                total += 1

            explore_ratios.append(explore_count / total if total > 0 else 0.0)
            avg_epsilons.append(np.mean(epsilons) if epsilons else 0.0)
            heatmap_data.append(row)

            if explore_count >= total / 2:
                explore_durations.append(episode["episode_duration"])
            else:
                exploit_durations.append(episode["episode_duration"])

        # 1. Explore ratio
        # plt.figure(figsize=(10, 5))
        # plt.plot(explore_ratios, label="Explore Ratio", color="blue")
        # plt.title("Explore Ratio Over Episodes")
        # plt.xlabel("Episode")
        # plt.ylabel("Explore Ratio")
        # plt.grid(True)
        # plt.tight_layout()
        # plt.savefig(os.path.join(output_dir, "explore_ratio_over_episodes.png"))
        # plt.close()

        # 2. Avg epsilon
        def create_config_label(config):
            import os
            from core.enums import Algorithm

            label = (
                f"Episodios: {config.episodes}\n"
                f"Max hops: {config.max_hops}\n"
                f"Timeout episodio: {config.episode_timeout_ms} ms\n"
                f"Prob. desconexión: {config.disconnection_probability}\n"
                f"Int. desconexión fija: {config.disconnection_interval_ms} ms\n"
                f"Int. reconexión fija: {config.reconnection_interval_ms} ms\n"
                f"Int. desconexión media: {config.mean_disconnection_interval_ms} ms\n"
                f"Int. reconexión media: {config.mean_reconnection_interval_ms} ms\n"
                f"Topología: {os.path.basename(config.topology_file)}\n"
                f"Secuencia de funciones: {' → '.join(f.value for f in config.functions_sequence)}"
            )

            if Algorithm.Q_ROUTING in config.algorithms:
                label += "\n"
                label += f"Usa epsilon decay: {'Sí' if config.use_epsilon_decay else 'No'}\n"
                if config.use_epsilon_decay:
                    label += (
                        f"Epsilon inicial: {config.initial_epsilon:.2f}\n"
                        f"Factor de decay: {config.epsilon_decay:.4f}\n"
                        f"Epsilon mínimo: {config.epsilon_min:.2f}"
                    )
                else:
                    label += f"Epsilon fijo: {config.fixed_epsilon:.2f}"

            return label

        plt.figure(figsize=(10, 5))
        plt.plot(avg_epsilons, label="Avg Epsilon", color="orange")
        plt.title("Epsilon Promedio sobre Episodio")
        plt.xlabel("Episodio")
        plt.ylabel("Epsilon")
        plt.grid(True)

        label = create_config_label(self.config)
        plt.annotate(label, xy=(1.01, 0), xycoords='axes fraction', fontsize=10,
                    ha='left', va='bottom', linespacing=1.5,
                    bbox=dict(facecolor='white', alpha=0.7, edgecolor='gray'))

        plt.tight_layout()
        plt.savefig(os.path.join(output_dir, "average_epsilon_over_episodes.png"), bbox_inches='tight')
        plt.close()

        # 3. Heatmap
        # max_len = max(len(row) for row in heatmap_data)
        # for row in heatmap_data:
        #     row.extend([0] * (max_len - len(row)))
        # heatmap_array = np.array(heatmap_data)
        #
        # plt.figure(figsize=(12, 6))
        # plt.imshow(heatmap_array.T, aspect="auto", cmap="coolwarm", interpolation="nearest")
        # plt.title("Explore/Exploit Heatmap (1 = Explore, 0 = Exploit)")
        # plt.xlabel("Episode")
        # plt.ylabel("Hop Number")
        # plt.colorbar(label="Decision")
        # plt.tight_layout()
        # plt.savefig(os.path.join(output_dir, "explore_exploit_heatmap.png"))
        # plt.close()

        # 4. Average delivery time by decision
        # labels = ['Explore-Dominant', 'Exploit-Dominant']
        # values = [
        #     np.mean(explore_durations) if explore_durations else 0,
        #     np.mean(exploit_durations) if exploit_durations else 0
        # ]
        #
        # plt.figure(figsize=(8, 5))
        # plt.bar(labels, values, color=["skyblue", "salmon"])
        # plt.title("Average Delivery Time by Dominant Policy")
        # plt.ylabel("Avg Episode Duration (ms)")
        # plt.tight_layout()
        # plt.savefig(os.path.join(output_dir, "avg_delivery_time_by_decision_type.png"))
        # plt.close()

reports_manager = ReportsManager()
